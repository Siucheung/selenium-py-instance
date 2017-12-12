from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException,InvalidSelectorException
import time,sys,datetime


class AutomationGUISetup():

    def setup(self):
        self.DataFile = 'automation_gui_data.xlsx'
        self.driver = webdriver.Firefox()
        self.workdata = load_workbook(filename = self.DataFile)
        self.Product_Data_Key = {
                'pagename' : 'Centurylink_C3000A_pages' ,
                'pageaddr' : 'Centurylink_C3000A_url' ,
                'elementname' : 'element_name' ,
                'elementloc' : 'element_location' ,
                'elementval' : 'element_value',
                }
        self.Product_Version = {
                'newverison' : '~/Taskspace/Firmwere/Centrylink_C3000A/*00g-*/*cfe*',
                'oldversion' : '',
        }

    def teardown(self):
        self.driver.close()

    def get_data_sheet(self,sheetname):
        workdata = self.workdata
        sheetobtain = workdata.get_sheet_by_name(sheetname)
        return sheetobtain

    def data_column_index(self,sheetname,guidetailname,valuename):
        Product_Data_Key = self.Product_Data_Key
        if sheetname == 'pages':
            pagesheet = self.get_data_sheet(sheetname)
            for i in range (1,pagesheet.max_row+1):
                for j in range (1,pagesheet.max_column+1):
                    if Product_Data_Key['pageaddr'] == pagesheet.cell(row=i,column=j).value:
                        datacolumnindex = j
                        return datacolumnindex#5
        elif sheetname == 'elements':
            datacolumnindex = self.is_elment_has_value(sheetname,guidetailname,valuename)
            return datacolumnindex
        else:
            datacolumnindex = 1
            return datacolumnindex

    def data_row_index(self,sheetname,guidetailname):
        Product_Data_Key = self.Product_Data_Key
        if sheetname == 'pages':
            pagesheet = self.get_data_sheet(sheetname)
            for i in range (1,pagesheet.max_row+1):#6
                for j in range (1,pagesheet.max_column+1):#5
                    if guidetailname == pagesheet.cell(row=i,column=j).value:
                        datarowindex = i
                        return datarowindex#2
        elif sheetname == 'elements':
            elementsheet = self.get_data_sheet(sheetname)
            for i in range (1,elementsheet.max_row+1):
                for j in range (1,elementsheet.max_column+1):
                    if guidetailname == elementsheet.cell(row=i,column=j).value:
                        datarowindex = i
                        return datarowindex
        else:
            datarowindex = 1
            return datarowindex

    def is_elment_has_value(self,sheetname,guidetailname,valuename):
        Product_Data_Key = self.Product_Data_Key
        elementsheet = self.get_data_sheet(sheetname)
        if valuename == '':
            for i in range (1,elementsheet.max_row+1):
                for j in range (1,elementsheet.max_column+1):
                    if Product_Data_Key['elementloc'] == elementsheet.cell(row=i,column=j).value:
                        datacolumnindex = j
                        return datacolumnindex
        else:
            for i in range (1,elementsheet.max_row+1):
                for j in range (1,elementsheet.max_column+1):
                    if Product_Data_Key['elementval'] == elementsheet.cell(row=i,column=j).value:
                        datacolumnindex = j
                        return datacolumnindex

    def gui_detail(self,sheetname,guidetailname,valuename = ''):
        elementsheet = self.get_data_sheet(sheetname)
        rowindex = self.data_row_index(sheetname,guidetailname)
        columnindex = self.data_column_index(sheetname,guidetailname,valuename)
        datadetail = elementsheet.cell(row=rowindex,column=columnindex).value
        return datadetail

    def element_find(self,gui_detail):
        driver = self.driver
        try:
            element_found = driver.find_element_by_id(gui_detail)
            return element_found
        #except e:
            #element_found = driver.find_element_by_xpath(gui_detail)
            #return element_found
        except NoSuchElementException:
            element_found = driver.find_element_by_xpath(gui_detail)
            return element_found

    def router_login(self):
        driver = self.driver
        home = self.gui_detail('pages','main_page',)
        element_find = self.element_find
        driver.get(home)
        try:
            logoutid = self.gui_detail('elements','mainpage_logout')
            element_find(logoutid).click()
        except:
            usernameid = self.gui_detail('elements','main_Username')
            user = self.gui_detail('elements','main_Username','user')
            passwordid = self.gui_detail('elements','main_Password')
            password = self.gui_detail('elements','main_Password','user')
            applyid = self.gui_detail('elements','main_Apply')
            element_find(usernameid).clear()
            element_find(passwordid).clear()
            element_find(usernameid).send_keys(user)
            element_find(passwordid).send_keys(password)
            element_find(applyid).click()
        time.sleep(5)

    def router_page_go(self,page_by_router):
        driver = self.driver
        pagesheet = self.get_data_sheet('pages')

        rowindex = self.data_row_index('pages',page_by_router)
        columnindex = self.data_column_index('pages',page_by_router,'')
        page = pagesheet.cell(row=rowindex,column=columnindex).value

        br0rowindex = self.data_row_index('pages','main_page')
        br0columnindex = self.data_column_index('pages','main_page','')
        br0 = pagesheet.cell(row=br0rowindex,column=br0columnindex).value

        driver.get(br0 + page + '.html')
        time.sleep(10)

    def router_firmware_upgrade(self):
        self.router_login()
        self.router_page_go('utilities_UpgradeFirmware')
        product_version = self.Product_Version
        element_find = self.element_find
        browserid = self.gui_detail('elements','uupgrade_Browse')
        upgradeid = self.gui_detail('elements','uupgrade_Upgrade')
        element_find(browserid).send_keys(product_version['newverison'])
        element_find(upgradeid).click()
        time.sleep(150)
        print '7777777'

    def router_reboot(self):
        self.router_login()
        self.router_page_go('utilities_home')
        rebootid = self.gui_detail('elements','ureboot_button')
        self.element_find(rebootid).click()
        time.sleep(150)

if __name__ == '__main__':
    test = AutomationGUISetup()
    test.setup()
    test.router_firmware_upgrade()
    test.teardown()
